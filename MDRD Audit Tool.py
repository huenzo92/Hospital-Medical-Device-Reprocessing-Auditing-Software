import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import string
from datetime import datetime

CURRENT_YEAR = str(datetime.now().year)

COUNT = 0

class EntryTopLevel(tk.Toplevel):

    def __init__(self, frame=None, *args, **kwargs):
        tk.Toplevel.__init__(self, *args, **kwargs)
        self.frame = frame
        self.handle_listbox_creation()
        self.wm_title('Entries Summary')

        try:
            self.iconbitmap('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\tray.ico')
        except:
            pass

    def get_df_longest_tray_name_length(self):

        longest = 0

        for row in range(self.frame.df.shape[0]):
            if len(self.frame.df.iloc[row,1]) > longest:
                longest = len(self.frame.df.iloc[row,1])

        return longest

    def handle_listbox_creation(self):

        self.frame_scrollbar = tk.Scrollbar(self, orient='vertical')
        self.frame_scrollbar.pack(side='right', fill='y')

        self.frame_scrollbar_horizontal = tk.Scrollbar(self, orient='horizontal')
        self.frame_scrollbar_horizontal.pack(side='bottom')

        self.top_lbox = tk.Listbox(self, font=('courier', 8), width=168, height=50, activestyle='none', yscrollcommand=self.frame_scrollbar.set, xscrollcommand=self.frame_scrollbar_horizontal.set)
        self.top_lbox.pack()

        self.frame_scrollbar.config(command=self.top_lbox.yview)
        self.frame_scrollbar_horizontal.config(command=self.top_lbox.xview)

        tray_name_longest_length = self.get_df_longest_tray_name_length()


        header_string = '        ' + 'DATE'.ljust(15) + 'ID'.ljust(16) + 'TRAY NAME'.ljust(tray_name_longest_length+5) + 'PREPARED BY'.ljust(30) + 'Q1'.ljust(8) + 'Q2'.ljust(8)
        header_string+= 'Q3'.ljust(8) + 'Q4'.ljust(8) + 'Q5'.ljust(8) + 'Q6'.ljust(8) + 'COMMENT'.ljust(100)

        header_gap_string = ''.ljust(17) + ''.ljust(16) + ''.ljust(tray_name_longest_length+5) + ''.ljust(30) + ''.ljust(8) + ''.ljust(8)
        header_gap_string+= ''.ljust(8) + ''.ljust(8) + ''.ljust(8) + ''.ljust(8) + ''.ljust(100)

        self.top_lbox.insert('end', header_string)
        self.top_lbox.insert('end', header_gap_string)

        index_number = 1

        for row in range(self.frame.df.shape[0]):
            row_string = ('(' + str(index_number) + ')').rjust(6); index_number+= 1
            row_string+= '  ' + str(self.frame.df.iloc[row,0]).ljust(15) # date
            row_string+= (str(self.frame.df.iloc[row,1].split('[')[1]).replace(']', ''))[:12].ljust(16) # ID
            row_string+= self.frame.df.iloc[row,1].ljust(tray_name_longest_length+5) # tray name
            row_string+= self.frame.df.iloc[row,2][:25].ljust(30) # prepared by
            row_string+= self.frame.df.iloc[row,3].ljust(8) # external wrap intact
            row_string+= self.frame.df.iloc[row,4].ljust(8) # tray identified
            row_string+= self.frame.df.iloc[row,5].ljust(8) # proper quantity sizing
            row_string+= self.frame.df.iloc[row,6].ljust(8) # chemical indicator included
            row_string+= self.frame.df.iloc[row,7].ljust(8) # instuments visually clean
            row_string+= self.frame.df.iloc[row,8].ljust(8) # instruments processed correctly
            row_string+= str(self.frame.df.iloc[row,9]).replace('\n', '     ').replace('nan', '')[:95].ljust(100) # comment

            self.top_lbox.insert('end', row_string)


class EntriesFrame(tk.Frame):

    def __init__(self, df=None, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)
        self.df = df

    def update_df(self, df):
        self.df = df

class SampleApp(tk.Tk):

    def __init__(self):

        tk.Tk.__init__(self)

        self.employee_list = ['Johns, Brianne (1)', 'Haner, Brian (2)', 'Carre, Vanessa (3)', 'Hedot, Brian (4)', 'Jimon, Fred (5)',
                              'Gregg, Samantha (6)', 'Dean, Billy (7)', 'Lyle, Kyle (8)', 'Degutye, Julia (9)', 'Marks, Matthew (10)']

        self.errors = {'date': 1,
                       'tray_name': 1,
                       'prepared_by': 1}

        # self.tray_df = pd.read_excel('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\Tray List\\Master Tray List.xlsx')
        # self.tray_list = list(self.tray_df['Tray Name'])

        self.initialize_tray_list()
#         self.prepare_dataframe()
        # self.handle_menu_bar_creation()
        self.handle_label_creation()
        self.handle_checkbox_creation()
        self.handle_entry_creation()
        self.handle_button_creation()
        self.handle_layout()

        # Window settings
        self.winfo_toplevel().title('MDRD Tray Audit Tool')
        self.top_level = None # This is to check later if a toplevel window already exists

        try:
            self.iconbitmap('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\tray.ico')
        except:
            pass
        self.resizable(False, False)

        # self.config(menu=self.menubar)

    def initialize_tray_list(self):

        self.tray_df = pd.read_excel('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\Tray List\\Master Tray List.xlsx')
        self.tray_list = list(self.tray_df['Tray Name'])

    def prepare_dataframe(self, employee_name):

        try:
            self.df = pd.read_excel('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Audit Data - ' + CURRENT_YEAR + ' - ' + employee_name + '.xlsx')

        except:
            # tk.messagebox.showinfo('Data Load Error', 'Cannot find master data file. Please contact Brendan.')
            self.df = pd.DataFrame(columns = ['Date', 'Tray Name', 'Prepared By', 'External Wrap Intact', 'Tray Identified', 'Proper Quantity/Sizing', 'Chemical Indicator Included', 'Instruments Visually Clean', 'Instruments Processed Correctly', 'Comment'])

    def handle_menu_bar_creation(self):

        self.menubar = tk.Menu(self)
        self.menubar.add_command(label='                            CLICK HERE TO VIEW ALL AUDIT ENTRIES                               ', command=self.create_window)

    def create_window(self):

        if self.top_level is None or not self.top_level.winfo_exists():
            self.prepare_dataframe()
            self.entries_summary = EntriesFrame(self.df, borderwidth=22)
            self.top_level = EntryTopLevel(frame=self.entries_summary)
            self.top_level.frame.update_df(self.df)
            self.top_level.wm_geometry("1200x762")
            self.top_level.resizable(False, False)

    def handle_label_creation(self):

        # CHECKBOXES
        self.q1 = tk.Label(self, font=('Calibri', 12), text="External Wrap Intact:")
        self.q2 = tk.Label(self, font=('Calibri', 12), text="Tray Identified:", pady=3)
        self.q3 = tk.Label(self, font=('Calibri', 12), text="Proper Quantity/Sizing:", pady=3)
        self.q4 = tk.Label(self, font=('Calibri', 12), text="Chemical Indicator Included:", pady=3)
        self.q5 = tk.Label(self, font=('Calibri', 12), text="Instruments Visually Clean:", pady=3)
        self.q6 = tk.Label(self, font=('Calibri', 12), text="Instruments Processed Correctly:", pady=3, padx=12)

        # 3 ENTRY FIELDS
        self.e1 = tk.Label(self, font=('Arial', 10), text="Date (dd/mm/yyyy): ", pady=3, padx=10)
        self.e2 = tk.Label(self, font=('Arial', 10), text="Prepared By: ", pady=3, padx=10)
        self.e3 = tk.Label(self, font=('Arial', 10), text="Tray Name: ", pady=3, padx=10)

        # COMMENT LABEL
        self.comment_label = tk.Label(self, font=('Arial', 11), text="Comment:")

        # ENTRY COUNT LABEL
        self.counter = tk.Label(self, font=('Calibri', 11), text="Entry Count: " + str(COUNT))

        # LAST ENTRY LABELS
        self.last_entry_header = tk.Label(self, font=('Calibri', 12, 'bold'), text="LAST ENTRY")
        self.last_date = tk.Label(self, font=('Calibri', 8), text = 'Date:')
        self.last_tray_name = tk.Label(self, font=('Calibri', 8), text = 'Tray ID:')
        self.last_prepared_by = tk.Label(self, font=('Calibri', 8), text= 'Prepared By:')
        self.last_q1 = tk.Label(self, font=('Calibri', 8), text = 'External Wrap Intact:')
        self.last_q2 = tk.Label(self, font=('Calibri', 8), text = 'Tray Identified:')
        self.last_q3 = tk.Label(self, font=('Calibri', 8), text = 'Proper Quantity/Sizing:')
        self.last_q4 = tk.Label(self, font=('Calibri', 8), text = 'Chemical Indicator Included:')
        self.last_q5 = tk.Label(self, font=('Calibri', 8), text = 'Instruments Visually Clean:')
        self.last_q6 = tk.Label(self, font=('Calibri', 8), text = 'Instruments Processed Correctly:')

        # BLANK BOTTOM ROW FOR SPACING
        self.bottom_spacer = tk.Label(self, text = ' ')

    def handle_entry_creation(self):

        # ENTRIES
        self.date = tk.Entry(self, width=10)
        self.date.bind('<FocusOut>', self.focus_date_validation)

        self.prepared_by_combobox = ttk.Combobox(self, width=34, values=self.employee_list)
        self.prepared_by_combobox.bind('<FocusOut>', self.focus_prepared_by_validation)

        # LISTBOX/SCROLLBAR
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_list) # Bind function to the variable in the tray name entry field

        self.tray_entry = tk.Entry(self, textvariable=self.search_var, width=37)

        self.scrollbar = tk.Scrollbar(self, orient='vertical')
        self.scrollbar.grid(row=10, column=2, sticky='ns')

        self.lbox = tk.Listbox(self, width=60, height=10, activestyle='none', yscrollcommand=self.scrollbar.set)
        self.lbox.bind('<FocusOut>', self.focus_tray_name_validation)

        self.scrollbar.config(command=self.lbox.yview)

        self.update_list() # Function for updating the list/doing the search. It needs to be called here to populate the listbox.

        # COMMENT
        self.comment = tk.Text(self, font=('Calibri', 9), height=3, width=55)

    def update_list(self, *args):

        self.tray_entry.config({'background': 'White'})
        self.lbox.delete(0, 'end')

        search_term = self.search_var.get().strip().lower()
        lbox_list = self.tray_list

        for item in lbox_list:
            if search_term in item.lower():
                self.lbox.insert('end', '  ' + item)

    def handle_checkbox_creation(self):

        self.check_1 = ttk.Checkbutton(self, text='Yes')
        self.check_2 = ttk.Checkbutton(self, text='Yes')
        self.check_3 = ttk.Checkbutton(self, text='Yes')
        self.check_4 = ttk.Checkbutton(self, text='Yes')
        self.check_5 = ttk.Checkbutton(self, text='Yes')
        self.check_6 = ttk.Checkbutton(self, text='Yes')

    def handle_button_creation(self):

        self.report_button = tk.Button(self, font=('Arial', 10), text=" Create Report ", command=self.create_report)
        self.submit_button = tk.Button(self, font=('Arial', 16, 'bold'), text="SUBMIT", command=self.on_button)

    def handle_layout(self):

        # CHECKBOXES AND THEIR LABELS
        self.q1.grid(row=0, column=0, sticky='', pady=(12,3)); self.check_1.grid(row=0, column=1, sticky='', pady=(12,3))
        self.check_1.state(['!alternate']); self.check_1.state(['selected']) # Remove 'alternate' state and set default to 'selected'

        self.q2.grid(row=1, column=0, sticky=''); self.check_2.grid(row=1, column=1, sticky='')
        self.check_2.state(['!alternate']); self.check_2.state(['selected'])

        self.q3.grid(row=2, column=0, sticky=''); self.check_3.grid(row=2, column=1, sticky='')
        self.check_3.state(['!alternate']); self.check_3.state(['selected'])

        self.q4.grid(row=3, column=0, sticky=''); self.check_4.grid(row=3, column=1, sticky='')
        self.check_4.state(['!alternate']); self.check_4.state(['selected'])

        self.q5.grid(row=4, column=0, sticky=''); self.check_5.grid(row=4, column=1, sticky='')
        self.check_5.state(['!alternate']); self.check_5.state(['selected'])

        self.q6.grid(row=5, column=0, sticky=''); self.check_6.grid(row=5, column=1, sticky='')
        self.check_6.state(['!alternate']); self.check_6.state(['selected'])

        # HORIZONTAL LINES GENERATION
        self.h_line = ttk.Separator(self, orient='horizontal')
        self.h_line_2 = ttk.Separator(self, orient='horizontal')
        self.h_line_3 = ttk.Separator(self, orient='horizontal')

        # HORIZONTAL LINE
        self.h_line.grid(row=6, column=0, sticky='we', columnspan=2, pady=10, padx=12)

        # 3 ENTRY FIELDS AND THEIR LABELS
        self.e1.grid(row=7, column=0, sticky='w', padx=5); self.date.grid(row=7, columnspan=2, padx=(5,15), sticky='e')
        self.e2.grid(row=8, column=0, sticky='w', padx=5); self.prepared_by_combobox.grid(row=8, columnspan=2, padx=(5,15), sticky='e')
        self.e3.grid(row=9, column=0, sticky='w', padx=5); self.tray_entry.grid(row=9, columnspan=2, padx=(5,15), pady=3, sticky='e')

        # LISTBOX
        self.lbox.grid(row=10, columnspan=2, padx=10, pady=10)

        # HORIZONTAL LINE
        self.h_line_2.grid(row=11, column=0, sticky='we', columnspan=2, pady=10, padx=12)

        # COMMENT BOX
        self.comment_label.grid(row=12, columnspan=2); self.comment.grid(row=13, columnspan=2, pady=10)

        # CREATE REPORT BUTTON + SUBMIT BUTTON + ENTRY COUNTER
        self.report_button.grid(row=14, column=0, sticky='w', padx=15, pady=(0,7))
        self.submit_button.grid(row=14, pady=(20,30), columnspan=2)
        self.counter.grid(row=14, column=1, padx=30, pady=(0,11))

################################################################################################################

    def focus_date_validation(self, event=None): # BIND METHOD ON A WIDGET PASSES THE EVENT OBJECT TO THE CALLBACK FUNCTION

        try:
            datetime.strptime(self.date.get(), '%d/%m/%Y') # Make sure the date is in the proper format

            if (datetime.strptime(self.date.get(), '%d/%m/%Y').year < 2019 or datetime.strptime(self.date.get(), '%d/%m/%Y').year > 2025):
                fail = 1/0

            self.date.config({"background": "Green"}) # If so, change the background color to green

        except:
            if self.date.get() != '':
                self.date.config({"background": "Red"})
            else:
                self.date.config({"background": "White"})

    def focus_prepared_by_validation(self, event=None):

        length = len(self.prepared_by_combobox.get().strip())

        employees_matched = []

        for employee in self.employee_list:
            if self.prepared_by_combobox.get().lower().strip() == employee[0:length].lower():
                employees_matched.append(employee)

        if len(employees_matched) == 1:
            self.prepared_by_combobox.set(employees_matched[0])

    def focus_tray_name_validation(self, event=None):

        if self.lbox.get('active').strip() in self.tray_list and len(self.lbox.curselection()) != 0:
            self.tray_to_delete = self.lbox.get('active').strip() # Because when you delete tray_entry text, the 'anchor' becomes ''
            self.tray_entry.delete(0, 'end') # This resets the 'anchor' text
            self.tray_entry.insert('end', ' ' + self.tray_to_delete)
            self.tray_entry.config({"background": "Green"})
        elif self.tray_entry.get().strip() in self.tray_list:
            self.tray_entry.config({"background": "Green"})

        elif self.tray_entry.get().strip() == '':
            self.tray_entry.config({"background": "White"})

        else:
            self.tray_entry.config({"background": "Red"})

################################################################################################################

    def date_validation(self, event=None): # BIND METHOD ON A WIDGET PASSES THE EVENT OBJECT TO THE CALLBACK FUNCTION

        try:
            datetime.strptime(self.date.get(), '%d/%m/%Y') # Make sure the date is in the proper format

            if (datetime.strptime(self.date.get(), '%d/%m/%Y').year < 2019 or datetime.strptime(self.date.get(), '%d/%m/%Y').year > 2025):
                fail = 1/0

            self.date.config({"background": "Green"}) # If so, change the background color to green
            self.errors['date'] = 0 # Alter the dictionary value to show that there's no error
        except:
            self.errors['date'] = 1
            self.date.config({"background": "Red"})

    def prepared_by_validation(self, event=None):

        length = len(self.prepared_by_combobox.get().strip())

        employees_matched = []

        for employee in self.employee_list:
            if self.prepared_by_combobox.get().lower().strip() == employee[0:length].lower():
                employees_matched.append(employee)

        if len(employees_matched) == 1:
            self.errors['prepared_by'] = 0
            self.prepared_by_combobox.set(employees_matched[0])
        else:
            self.errors['prepared_by'] = 1
            if self.prepared_by_combobox.get() != '':
                tk.messagebox.showinfo('Input Error', 'Please enter a valid entry into \'Prepared By:\'')

    def tray_name_validation(self, event=None):
        if self.lbox.get('active').strip() in self.tray_list and len(self.lbox.curselection()) != 0:
            self.tray_to_delete = self.lbox.get('active').strip() # Because when you delete tray_entry text, the 'anchor' becomes ''
            self.tray_entry.delete(0, 'end') # This resets the 'anchor' text
            self.tray_entry.insert('end', ' ' + self.tray_to_delete)
            self.tray_entry.config({"background": "Green"})
            self.errors['tray_name'] = 0

        elif self.tray_entry.get().strip() in self.tray_list:
            self.tray_entry.config({"background": "Green"})
            self.errors['tray_name'] = 0

        else:
            self.tray_entry.config({"background": "Red"})
            self.errors['tray_name'] = 1

################################################################################################################

    def get_checkbox_answers(self):

        self.q1_answer = 'No'; self.q2_answer = 'No'; self.q3_answer = 'No'; self.q4_answer = 'No'; self.q5_answer = 'no'; self.q6_answer = 'No'

        try:
            if (self.check_1.state()[0] == 'selected' or self.check_1.state()[1] == 'selected'): # Output will be one of: (), ('focus'), ('selected'), or ('focus', 'selected')
                self.q1_answer = 'Yes'
        except:
            pass

        try:
            if (self.check_2.state()[0] == 'selected' or self.check_2.state()[1] == 'selected'):
                self.q2_answer = 'Yes'
        except:
            pass

        try:
            if (self.check_3.state()[0] == 'selected' or self.check_3.state()[1] == 'selected'):
                self.q3_answer = 'Yes'
        except:
            pass

        try:
            if (self.check_4.state()[0] == 'selected' or self.check_4.state()[1] == 'selected'):
                self.q4_answer = 'Yes'
        except:
            pass

        try:
            if (self.check_5.state()[0] == 'selected' or self.check_5.state()[1] == 'selected'):
                self.q5_answer = 'Yes'
        except:
            pass

        try:
            if (self.check_6.state()[0] == 'selected' or self.check_6.state()[1] == 'selected'):
                self.q6_answer = 'Yes'
        except:
            pass

    def append_row_to_df(self):

        self.row_to_append = {'Date': self.date.get(),
                              'Tray Name': self.tray_entry.get(),
                              'Prepared By': self.prepared_by_combobox.get().strip(),
                              'External Wrap Intact': self.q1_answer,
                              'Tray Identified': self.q2_answer,
                              'Proper Quantity/Sizing': self.q3_answer,
                              'Chemical Indicator Included': self.q4_answer,
                              'Instruments Visually Clean': self.q5_answer,
                              'Instruments Processed Correctly': self.q6_answer,
                              'Comment': self.comment.get('1.0', 'end-1c')
                             }

        self.df = self.df.append(self.row_to_append, ignore_index=True)

    def increment_counter(self):

        global COUNT
        COUNT = COUNT + 1
        self.counter = tk.Label(self, font=('Calibri', 11), text="Entry Count: " + str(COUNT))
        self.counter.grid(row=14, column=1, padx=30, pady=(0,11))

    def display_last_entry(self):

        # HORIZONTAL LINE
        self.h_line_3.grid(row=15, column=0, sticky='we', columnspan=2, pady=(0,15), padx=12)

        # Place labels
        self.last_entry_header.grid(row=16, columnspan=2, sticky='')
        self.last_date.grid(row=17, column=0, sticky='')
        self.last_tray_name.grid(row=18, column=0, sticky='')
        self.last_prepared_by.grid(row=19, column=0, sticky='')
        self.last_q1.grid(row=20, column=0, sticky='')
        self.last_q2.grid(row=21, column=0, sticky='')
        self.last_q3.grid(row=22, column=0, sticky='')
        self.last_q4.grid(row=23, column=0, sticky='')
        self.last_q5.grid(row=24, column=0, sticky='')
        self.last_q6.grid(row=25, column=0, sticky='')

        # Get the last row of the dataframe, then extract the values...
        self.last_date_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [0]))
        self.last_date_answer.grid(row=17, column=1)

        self.last_tray_name_answer = tk.Label(self, font=('Calibri', 8), text='[' + str(self.df.iloc[self.df.shape[0]-1] [1].split('[')[1]))
        self.last_tray_name_answer.grid(row=18, column=1)

        self.last_prepared_by_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [2]))
        self.last_prepared_by_answer.grid(row=19, column=1)

        self.last_q1_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [3]))
        self.last_q1_answer.grid(row=20, column=1)

        self.last_q2_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [4]))
        self.last_q2_answer.grid(row=21, column=1)

        self.last_q3_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [5]))
        self.last_q3_answer.grid(row=22, column=1)

        self.last_q4_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [6]))
        self.last_q4_answer.grid(row=23, column=1)

        self.last_q5_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [7]))
        self.last_q5_answer.grid(row=24, column=1)

        self.last_q6_answer = tk.Label(self, font=('Calibri', 8), text=str(self.df.iloc[self.df.shape[0]-1] [8]))
        self.last_q6_answer.grid(row=25, column=1)

        # Blank label just to add space at very bottom
        self.bottom_spacer.grid(row=26)

    def save_files(self, employee_name):

        workbook = Workbook()
        worksheet = workbook.worksheets[0]
        worksheet.title = 'Audit Data'

        rows = dataframe_to_rows(self.df, index=False)

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 60
        worksheet.column_dimensions['C'].width = 32
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 30
        worksheet.column_dimensions['F'].width = 30
        worksheet.column_dimensions['G'].width = 30
        worksheet.column_dimensions['H'].width = 30
        worksheet.column_dimensions['I'].width = 30
        worksheet.column_dimensions['J'].width = 60

        for row_index, row in enumerate(rows, 1):
            for column_index, value in enumerate(row, 1):
                 worksheet.cell(row=row_index, column=column_index, value=value).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        self.timestamp = str(datetime.now().date()) + ' ' + str(datetime.now().hour) + '-' + str(datetime.now().minute) + '-' + str(datetime.now().second)

        workbook.save('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Audit Data - ' + CURRENT_YEAR + ' - ' + employee_name + '.xlsx')
        workbook.save('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\Audit Data Backups\\Audit Data - ' + self.timestamp + ' - ' + employee_name + '.xlsx')

    def update_tray_list(self):

        index_to_drop = self.tray_list.index(self.tray_to_delete)
        self.tray_df.drop(index_to_drop, inplace=True)
        self.tray_df.reset_index(drop=True, inplace=True)
        self.tray_list = list(self.tray_df['Tray Name'])
        self.update_list() # Dynamically remove the submission from the listbox

        workbook = Workbook()
        worksheet = workbook.worksheets[0]
        worksheet.title = 'Tray List'

        rows = dataframe_to_rows(self.tray_df, index=False)

        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 30

        for row_index, row in enumerate(rows, 1):
            for column_index, value in enumerate(row, 1):
                 worksheet.cell(row=row_index, column=column_index, value=value).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        workbook.save('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\Tray List.xlsx')

    def reset_checkboxes(self):

        self.check_1.state(['selected'])
        self.check_2.state(['selected'])
        self.check_3.state(['selected'])
        self.check_4.state(['selected'])
        self.check_5.state(['selected'])
        self.check_6.state(['selected'])

    def reset_entries(self):

        self.date.delete(0, 'end') # Clear entry field text
        self.date.config({"background": "White"}) # Reset background color

        self.prepared_by_combobox.delete(0, 'end')

        self.tray_entry.delete(0, 'end')
        self.tray_entry.config({"background": "White"})

        self.comment.delete('1.0', 'end')

    def on_button(self):

        # Run validation functions
        self.date_validation()
        self.prepared_by_validation()
        self.tray_name_validation()

        # If none of the three Entry widgets have a data validation error...
        if sum(self.errors.values()) == 0:

            self.initialize_tray_list()
            # self.update_tray_list()
            self.prepare_dataframe(self.prepared_by_combobox.get().strip())
            self.get_checkbox_answers()
            self.append_row_to_df()
            self.increment_counter()
            self.display_last_entry()
            self.save_files(self.prepared_by_combobox.get().strip())
            self.reset_checkboxes()
            self.reset_entries()

            # Update the Frame df. Use exception handling because if the frame hasn't been created, it will be a None object
            try:
                self.top_level.frame.update_df(self.df)
                 # Remove listbox from grid and recreate
                self.top_level.top_lbox.pack_forget()
                self.top_level.frame_scrollbar.pack_forget()
                self.top_level.handle_listbox_creation()
            except:
                pass

    def create_report(self):

        for employee_name in self.employee_list:

            # Empty dataframe to append to and export
            self.report_df = pd.DataFrame(columns = ['File', 'Tray Name', 'ID'])

            try:
                # Import currently audited trays, and put into list of tray names
                self.imported_audit_df = pd.read_excel('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Audit Data - ' + CURRENT_YEAR + ' - ' + employee_name + '.xlsx')
                self.audited_list = list(self.imported_audit_df['Tray Name'])
                self.audited_list = [x.strip() for x in self.audited_list]

                self.master_tray_df = pd.read_excel('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Backups\\Tray List\\Master Tray List.xlsx')

                # Go through every tray. If in audited list, append two columns to the to-be-exported dataframe
                for row in range(self.master_tray_df.shape[0]):
                    if self.master_tray_df.iloc[row]['Tray Name'] in self.audited_list:
                        pass
                    else:
                        self.report_df = self.report_df.append(self.master_tray_df.loc[row, ['File', 'Tray Name', 'ID']])

                workbook_report = Workbook()
                worksheet_report = workbook_report.worksheets[0]
                worksheet_report.title = 'Audit Report'

                rows_report = dataframe_to_rows(self.report_df, index=False)

                worksheet_report.column_dimensions['A'].width = 35
                worksheet_report.column_dimensions['B'].width = 100
                worksheet_report.column_dimensions['C'].width = 20

                for row_index, row in enumerate(rows_report, 1):
                    for column_index, value in enumerate(row, 1):
                         worksheet_report.cell(row=row_index, column=column_index, value=value).alignment = Alignment(horizontal = 'center')

                workbook_report.save('L:\\Facilities\\Common\\MDRD\\CONFIDENTIAL\\Audit tracking MDRD\\AUDIT INPUT TOOL\\Audit Report - ' + CURRENT_YEAR + ' - ' + employee_name + '.xlsx')

            except:
                pass

        tk.messagebox.showinfo('Success', 'Report Generated!')

app = SampleApp()
app.mainloop()


# In[ ]:
